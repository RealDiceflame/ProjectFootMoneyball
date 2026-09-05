from openpyxl import load_workbook
import pandas as pd
from app.draft_board_exporter import BOARD_COLUMNS, export_switchable_draft_board


def test_exporter_builds_switchable_workbook(tmp_path):
    row = {column: 1 for column in BOARD_COLUMNS}
    row.update({"player": "Test Player", "team": "BUF", "pos": "QB", "format": "test"})
    for teams in (8, 10, 12, 14, 16):
        for index in range(12):
            pd.DataFrame([row]).to_csv(tmp_path / f"draft_rankings_{teams}team_format_{index}.csv", index=False)
    destination = tmp_path / "board.xlsx"
    export_switchable_draft_board(tmp_path, destination)
    workbook = load_workbook(destination, data_only=False)
    assert workbook.sheetnames == ["Draft Board", "League Settings", "Format Data"]
    assert workbook["Draft Board"]["A1"].value == "OutlierBaseline Fantasy Draft Board"
    assert workbook["Draft Board"]["V2"].value == 12
    assert workbook["Draft Board"]["Z2"].value == "2QB"
    assert workbook["Draft Board"]["V3"].value == "Half PPR"
    assert workbook["Draft Board"]["Z3"].value == "+0.5"
    assert workbook["Draft Board"]["A5"].value.startswith("=SORT(FILTER(")
    assert 'I5>=50' in workbook["Draft Board"]["S5"].value
    assert 'I5="","NO MARKET"' in workbook["Draft Board"]["S5"].value
    assert 'I5>=25' in workbook["Draft Board"]["S5"].value
    assert 'I5<=-20' in workbook["Draft Board"]["S5"].value
    assert workbook["Format Data"].max_row == 61
