"""Export ranking CSVs as one switchable draft-board workbook."""

from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.draft_tags import REACH_MAX, TARGET_MIN, VALUE_MIN

TEAM_SIZES = (8, 10, 12, 14, 16)
BOARD_COLUMNS = ["overall_rank", "player", "team", "pos", "position_rank",
                 "projected_points", "replacement_points", "vorp", "market_value", "adp",
                 "source_count", "adp_spread", "value_vs_adp", "Yahoo", "Sleeper", "NFL",
                 "FFC", "MFL", "format"]


def export_switchable_draft_board(rankings_dir, workbook_path):
    """Create an Excel/Google Sheets-ready board with four setting dropdowns."""
    rankings_dir = Path(rankings_dir)
    files = []
    for teams in TEAM_SIZES:
        files.extend(sorted(rankings_dir.glob(f"draft_rankings_{teams}team_*.csv")))
    if len(files) != 60:
        raise FileNotFoundError(
            f"Expected 60 team-size ranking CSVs in {rankings_dir}, found {len(files)}."
        )
    combined = pd.concat([pd.read_csv(path) for path in files], ignore_index=True)
    combined = combined[BOARD_COLUMNS]
    workbook_path = Path(workbook_path)
    workbook_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    board = wb.active
    board.title = "Draft Board"
    settings = wb.create_sheet("League Settings")
    data = wb.create_sheet("Format Data")
    navy, blue, white = "17233B", "2F75B5", "FFFFFF"

    data.append(BOARD_COLUMNS)
    for row in combined.itertuples(index=False, name=None):
        data.append(list(row))
    for cell in data[1]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color=white, bold=True)
    data.freeze_panes = "A2"
    data.auto_filter.ref = data.dimensions

    board.merge_cells("A1:T1")
    board["A1"] = "OutlierBaseline Fantasy Draft Board"
    board["A1"].fill = PatternFill("solid", fgColor=navy)
    board["A1"].font = Font(color=white, bold=True, size=18)
    board["A1"].alignment = Alignment(horizontal="center")
    board.merge_cells("A2:T2")
    board["A2"] = '=V2&" teams | "&Z2&" | "&V3&" | TE premium "&Z3'
    board["A2"].alignment = Alignment(horizontal="center")
    headers = ["Rank", "Player", "Team", "Pos", "Pos Rank", "Projected Pts",
               "Replacement Pts", "VORP", "Market +/-", "ADP", "Sources",
               "ADP Spread", "Value vs ADP", "Yahoo", "Sleeper", "NFL/ESPN", "FFC",
               "MFL", "Format", "Draft Tag"]
    for column, header in enumerate(headers, 1):
        cell = board.cell(4, column, header)
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.font = Font(color=white, bold=True)

    last_row = data.max_row
    format_formula = ('$V$2&"-team "&IF($Z$2="2QB","2QB ","1QB ")&'
                      'IF($V$3="Standard","standard",IF($V$3="Full PPR","full-PPR","half-PPR"))&'
                      'IF($Z$3="+0.5"," + 0.5 TE premium","")')
    board["A5"] = (f"=SORT(FILTER('Format Data'!A2:S{last_row},"
                   f"'Format Data'!S2:S{last_row}=({format_formula})),1,TRUE)")
    for row in range(5, 205):
        board.cell(
            row,
            20,
            f'=IF(A{row}="","",IF(I{row}>={TARGET_MIN:g},"TARGET",'
            f'IF(I{row}>={VALUE_MIN:g},"VALUE",IF(I{row}<={REACH_MAX:g},"REACH","FAIR"))))',
        )
        if row % 2 == 0:
            for column in range(1, 21):
                board.cell(row, column).fill = PatternFill("solid", fgColor="F3F6FA")

    board.merge_cells("U1:AB1")
    board["U1"] = "Ranking Controls"
    board["U1"].fill = PatternFill("solid", fgColor=navy)
    board["U1"].font = Font(color=white, bold=True)
    board["U1"].alignment = Alignment(horizontal="center")
    board["U2"], board["V2"], board["Y2"], board["Z2"] = "Teams", 12, "QB", "2QB"
    board["U3"], board["V3"], board["Y3"], board["Z3"] = "PPR", "Half PPR", "TE Premium", "+0.5"
    for range_ref in ("V2:X2", "Z2:AB2", "V3:X3", "Z3:AB3"):
        board.merge_cells(range_ref)
    for cell_ref in ("U2", "Y2", "U3", "Y3"):
        board[cell_ref].fill = PatternFill("solid", fgColor="D9EAF7")
        board[cell_ref].font = Font(bold=True)
    for cell_ref, choices in (("V2", '"8,10,12,14,16"'), ("Z2", '"1QB,2QB"'),
                              ("V3", '"Standard,Half PPR,Full PPR"'), ("Z3", '"Off,+0.5"')):
        validation = DataValidation(type="list", formula1=choices)
        validation.error = "Choose a value from the dropdown."
        validation.showErrorMessage = True
        board.add_data_validation(validation)
        validation.add(board[cell_ref])

    for range_ref, operator, formulas, color in (
        ("I5:I204", "greaterThanOrEqual", [f"{TARGET_MIN:g}"], "C6EFCE"),
        ("I5:I204", "between", [f"{VALUE_MIN:g}", f"{TARGET_MIN - 0.001:g}"], "FFEB9C"),
        ("I5:I204", "lessThanOrEqual", [f"{REACH_MAX:g}"], "FFC7CE")):
        board.conditional_formatting.add(range_ref, CellIsRule(operator=operator, formula=formulas,
            fill=PatternFill("solid", fgColor=color)))
    board.freeze_panes = "A5"
    board.auto_filter.ref = "A4:T204"
    for index, width in enumerate([8, 24, 8, 7, 10, 14, 16, 10, 12, 9, 8, 10, 13, 9, 9, 10, 9, 9, 35, 11], 1):
        board.column_dimensions[get_column_letter(index)].width = width

    settings.append(["League Setting", "Selected Value"])
    settings.append(["Teams", "='Draft Board'!V2"])
    settings.append(["Quarterbacks", "='Draft Board'!Z2"])
    settings.append(["Reception scoring", "='Draft Board'!V3"])
    settings.append(["TE premium", "='Draft Board'!Z3"])
    settings.append([])
    settings.append(["How to use", "Change the four dropdowns on Draft Board. Rankings update automatically."])
    for cell in settings[1]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color=white, bold=True)
    settings.column_dimensions["A"].width = 24
    settings.column_dimensions["B"].width = 78
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(workbook_path)
    return workbook_path
